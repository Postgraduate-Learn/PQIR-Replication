from lxml import etree
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.hyperlink import Hyperlink
from conf import conf
from base_process.util.split_body_text_and_code import extract_body_and_code

def writerObjs2xml(xml_file, objs, obj_name):
    """将数据写入到xml文件进行保存"""
    root = etree.Element(obj_name, attrib={'count': str(len(objs))})

    for obj in objs:
        try:
            etree.SubElement(root, 'row', attrib=obj.to_dict_s())
        except Exception as e:
            print('***** Error in writeObjs2xml() *****', e, ':', obj.to_dict_s())

    tree = etree.ElementTree(root)
    tree.write(xml_file, pretty_print=True, xml_declaration=True, encoding='utf-8')


def writer_file_excel(sample_excel_path, objs):
    """将数据写入到excel进行保存"""
    data = {
        'Index': [],
        'Title': [],
        'Score': [],
        'Tag': [],
        'Question_url': [],
        'IsAcceptedAnswer': [],
        'Category': [],
        'Body': [],
        'Code_In_Body': []
    }

    for count, obj in enumerate(objs, start=1):
        data['Index'].append(count)
        data['Title'].append(obj.title)
        data['Score'].append(obj.score)

        tag_list = obj.tags.strip('|').split('|') if obj.tags else []
        data['Tag'].append(tag_list)

        data['Question_url'].append(conf.so_url + '/questions/' + obj.id)
        if obj.accepted_answer_id is not None:
            data['IsAcceptedAnswer'].append('Y')
        else:
            data['IsAcceptedAnswer'].append('N')
        data['Category'].append("               ")
        cleaned_body, extracted_code = extract_body_and_code(obj.body)
        data['Body'].append(cleaned_body)
        data['Code_In_Body'].append(extracted_code)

    df = pd.DataFrame(data)
    df.to_excel(sample_excel_path, index=False)

    # 打开保存的 Excel 文件
    wb = load_workbook(sample_excel_path)
    ws = wb.active

    # 自适应列宽并设置对齐
    for col_num, column_cells in enumerate(ws.columns, 1):
        adjusted_width = 50
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # 设置单元格对齐，并启用换行和缩进，增加美观度
        for cell in column_cells:
            cell.alignment = Alignment(
                # 水平居中
                horizontal="center",
                # 垂直居中
                vertical="center",
                # 自动换行
                wrap_text=True
            )

    # 将 Question_url 列的内容变为超链接
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            # 获取链接
            url = cell.value
            if url:
                # 设置超链接
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
                # 应用超链接样式
                cell.style = "Hyperlink"

    # 调整行高
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 25

    # 保存修改后的 Excel 文件
    wb.save(sample_excel_path)

    print(f'成功写入{len(objs)}条记录到{sample_excel_path}')
